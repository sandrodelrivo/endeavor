"""Tkinter GUI for editing tier_templates.yaml + biome_overrides.yaml + tag JSONs.

Three panes:
    Left   - tiers (from tier_templates.yaml) + tags (from endeavour tag JSONs)
    Middle - features (typically ores) assigned to the selected tier/tag
    Right  - biomes assigned to the selected tier/tag

Buttons let you add/remove tiers, tags, ores, and biomes; the menu bar saves
all four backing files (tier_templates.yaml, biome_overrides.yaml, tag JSONs,
tier-map.xlsx) and runs apply.py to regenerate the biome JSONs.

Run from this directory:
    python gui.py
"""

from __future__ import annotations

import json
import subprocess
import sys
import tkinter as tk
from collections import OrderedDict
from pathlib import Path
from tkinter import messagebox, simpledialog, ttk

import openpyxl
import yaml

from common import (
    CONFIG_DIR,
    DATAPACK_ROOT,
    ENDEAVOUR_TAG_DIR,
    GENERATION_STEPS,
    OP_ADD,
    OP_INHERIT,
    OP_REMOVE,
    OP_SET,
    TIER_MAP_XLSX,
    TOOL_ROOT,
    all_biome_files,
    biome_id_from_path,
    normalize_biome_id,
    yaml_dump,
    yaml_load,
)


TT_PATH = CONFIG_DIR / "tier_templates.yaml"
BO_PATH = CONFIG_DIR / "biome_overrides.yaml"

# Default generation step we put a feature into when the user doesn't pick one.
DEFAULT_STEP = "underground_ores"


# --- Data model -----------------------------------------------------------

class DataModel:
    """Loads + mutates the four backing stores; saves them back atomically.

    Edits stay in memory until save_all() runs. apply.py is invoked
    separately so the user can sanity-check the YAML diff first.
    """

    def __init__(self) -> None:
        self.load()

    # --- load ---

    def load(self) -> None:
        self.biome_files = {biome_id_from_path(f): f for f in all_biome_files()}
        self.all_biome_ids = sorted(self.biome_files)

        self._tt_header = _read_header(TT_PATH)
        self._bo_header = _read_header(BO_PATH)
        self.tier_templates = (yaml_load(TT_PATH).get("tiers") or {})
        self.biome_overrides = (yaml_load(BO_PATH) or {})

        self.tags = self._load_tags()  # {tag_name: {"replace": bool, "values": [biome_id,...]}}
        self.tier_map, self._wb = self._load_xlsx_tiers()  # {biome_id: tier_str}

        self.known_features = self._collect_known_features()

    def _load_tags(self) -> dict[str, dict]:
        out: dict[str, dict] = {}
        for f in sorted(ENDEAVOUR_TAG_DIR.glob("*.json")):
            data = json.loads(f.read_text(encoding="utf-8"))
            values: list[str] = []
            for v in data.get("values", []):
                if isinstance(v, str):
                    values.append(v)
                elif isinstance(v, dict) and "id" in v:
                    values.append(v["id"])
            out[f.stem] = {
                "replace": bool(data.get("replace", False)),
                "values": values,
                "_path": f,
            }
        return out

    def _load_xlsx_tiers(self):
        wb = openpyxl.load_workbook(TIER_MAP_XLSX)
        ws = wb["Biomes"]
        out: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2):
            if row[0].value is None:
                continue
            bid = normalize_biome_id(str(row[0].value).strip(), set(self.biome_files))
            out[bid] = (str(row[4].value).strip() if row[4].value else "")
        return out, wb

    def _collect_known_features(self) -> list[str]:
        feats: set[str] = set()
        for f in self.biome_files.values():
            data = json.loads(f.read_text(encoding="utf-8"))
            for step in data.get("features", []):
                for x in step:
                    feats.add(x)
        _gather_features(self.tier_templates, feats)
        _gather_features(self.biome_overrides, feats)
        return sorted(feats)

    # --- queries used by the GUI ---

    def list_tiers(self) -> list[str]:
        # Tier_templates can be missing entries; merge with whatever the xlsx
        # already references so a tier with no template still shows up.
        tiers = set(self.tier_templates) | {t for t in self.tier_map.values() if t}
        # Sort with TURN OFF / TX last
        ordering = {"T1": 0, "T2": 1, "T3": 2, "T4": 3, "T5": 4, "TX": 8, "TURN OFF": 9}
        return sorted(tiers, key=lambda t: (ordering.get(t, 5), t))

    def list_tags(self) -> list[str]:
        return sorted(self.tags)

    def features_for_tier(self, tier: str, step: str = DEFAULT_STEP) -> list[str]:
        block = (self.tier_templates.get(tier) or {}).get(step)
        if block is None:
            return []
        if isinstance(block, list):
            return list(block)
        out: list[str] = []
        if OP_INHERIT in block:
            out = self.features_for_tier(block[OP_INHERIT], step)
        if OP_SET in block:
            out = list(block[OP_SET])
        for f in block.get(OP_ADD, []):
            if f not in out:
                out.append(f)
        for f in block.get(OP_REMOVE, []):
            if f in out:
                out.remove(f)
        return out

    def features_for_tag(self, tag: str, step: str = DEFAULT_STEP) -> list[str]:
        key = f"#endeavour:{tag}"
        block = (self.biome_overrides.get(key) or {}).get(step) or {}
        if isinstance(block, list):
            return list(block)
        # We use @add for tag-level injection in this tool.
        return list(block.get(OP_ADD, []))

    def biomes_for_tier(self, tier: str) -> list[str]:
        return sorted(b for b, t in self.tier_map.items() if t == tier)

    def biomes_for_tag(self, tag: str) -> list[str]:
        return sorted(self.tags.get(tag, {}).get("values", []))

    def tier_of(self, biome: str) -> str:
        return self.tier_map.get(biome, "")

    # --- mutations ---

    def add_tier(self, tier: str) -> None:
        tier = tier.strip()
        if not tier:
            raise ValueError("tier name required")
        if tier in self.tier_templates:
            raise ValueError(f"tier {tier!r} already exists")
        self.tier_templates[tier] = {DEFAULT_STEP: []}

    def add_tag(self, tag: str) -> None:
        tag = tag.strip().lstrip("#").removeprefix("endeavour:")
        if not tag:
            raise ValueError("tag name required")
        if tag in self.tags:
            raise ValueError(f"tag {tag!r} already exists")
        self.tags[tag] = {"replace": False, "values": [], "_path": None}

    def delete_tier(self, tier: str) -> None:
        if tier in self.tier_templates:
            del self.tier_templates[tier]
        # Don't clear xlsx tier assignments - that would be destructive without
        # confirmation. Tell the user to reassign biomes manually if needed.

    def delete_tag(self, tag: str) -> None:
        if tag in self.tags:
            # Mark the file for deletion on save by storing a sentinel.
            self.tags[tag]["_delete"] = True
            self.tags[tag]["values"] = []
        key = f"#endeavour:{tag}"
        if key in self.biome_overrides:
            del self.biome_overrides[key]

    def add_feature_to_tier(self, tier: str, feature: str,
                            step: str = DEFAULT_STEP) -> None:
        tt = self.tier_templates.setdefault(tier, {})
        block = tt.get(step)
        if block is None or isinstance(block, list):
            lst = list(block) if isinstance(block, list) else []
            if feature not in lst:
                lst.append(feature)
            tt[step] = lst
            return
        # dict form: prefer adding to @set if present, else @add
        if OP_SET in block:
            if feature not in block[OP_SET]:
                block[OP_SET].append(feature)
        else:
            adds = block.setdefault(OP_ADD, [])
            if feature not in adds:
                adds.append(feature)

    def remove_feature_from_tier(self, tier: str, feature: str,
                                 step: str = DEFAULT_STEP) -> None:
        tt = self.tier_templates.get(tier) or {}
        block = tt.get(step)
        if block is None:
            return
        if isinstance(block, list):
            if feature in block:
                block.remove(feature)
            return
        # dict form: remove from @add or @set; if it was inherited, add to @remove
        removed = False
        if OP_ADD in block and feature in block[OP_ADD]:
            block[OP_ADD].remove(feature)
            removed = True
        if OP_SET in block and feature in block[OP_SET]:
            block[OP_SET].remove(feature)
            removed = True
        if not removed and OP_INHERIT in block:
            block.setdefault(OP_REMOVE, [])
            if feature not in block[OP_REMOVE]:
                block[OP_REMOVE].append(feature)

    def add_feature_to_tag(self, tag: str, feature: str,
                           step: str = DEFAULT_STEP) -> None:
        key = f"#endeavour:{tag}"
        ov = self.biome_overrides.setdefault(key, {})
        block = ov.setdefault(step, {})
        if isinstance(block, list):
            if feature not in block:
                block.append(feature)
            return
        adds = block.setdefault(OP_ADD, [])
        if feature not in adds:
            adds.append(feature)

    def remove_feature_from_tag(self, tag: str, feature: str,
                                step: str = DEFAULT_STEP) -> None:
        key = f"#endeavour:{tag}"
        ov = self.biome_overrides.get(key) or {}
        block = ov.get(step)
        if block is None:
            return
        if isinstance(block, list):
            if feature in block:
                block.remove(feature)
            if not block:
                del ov[step]
            return
        if OP_ADD in block and feature in block[OP_ADD]:
            block[OP_ADD].remove(feature)
            if not block[OP_ADD]:
                del block[OP_ADD]
        if not block:
            del ov[step]
        if not ov:
            del self.biome_overrides[key]

    def assign_biome_to_tier(self, biome: str, tier: str) -> None:
        # The xlsx is the source of truth for tier mapping. A biome has one
        # tier; setting a new tier replaces whatever was there.
        if biome not in self.biome_files:
            raise ValueError(f"unknown biome {biome!r}")
        self.tier_map[biome] = tier

    def add_biome_to_tag(self, biome: str, tag: str) -> None:
        if biome not in self.biome_files:
            raise ValueError(f"unknown biome {biome!r}")
        if tag not in self.tags:
            raise ValueError(f"unknown tag {tag!r}")
        vals = self.tags[tag]["values"]
        if biome not in vals:
            vals.append(biome)

    def remove_biome_from_tag(self, biome: str, tag: str) -> None:
        if tag not in self.tags:
            return
        vals = self.tags[tag]["values"]
        if biome in vals:
            vals.remove(biome)

    # --- save ---

    def save_all(self) -> None:
        self._save_tier_templates()
        self._save_biome_overrides()
        self._save_tags()
        self._save_xlsx()

    def _save_tier_templates(self) -> None:
        # Preserve key order (python 3.7+ dicts), drop empty step dicts
        tiers_clean: dict[str, dict] = {}
        for tier, steps in self.tier_templates.items():
            clean_steps = {s: v for s, v in steps.items() if not _is_empty_step(v)}
            if clean_steps:
                tiers_clean[tier] = clean_steps
        yaml_dump(TT_PATH, {"tiers": tiers_clean}, header=self._tt_header)

    def _save_biome_overrides(self) -> None:
        # Drop empty entries
        bo_clean: dict[str, dict] = {}
        for key, steps in self.biome_overrides.items():
            if not steps:
                continue
            clean_steps = {s: v for s, v in steps.items() if not _is_empty_step(v)}
            if clean_steps:
                bo_clean[key] = clean_steps
        yaml_dump(BO_PATH, bo_clean, header=self._bo_header)

    def _save_tags(self) -> None:
        ENDEAVOUR_TAG_DIR.mkdir(parents=True, exist_ok=True)
        for tag, info in list(self.tags.items()):
            path: Path = info.get("_path") or (ENDEAVOUR_TAG_DIR / f"{tag}.json")
            if info.get("_delete"):
                if path.exists():
                    path.unlink()
                del self.tags[tag]
                continue
            data = {
                "replace": info.get("replace", False),
                "values": sorted(set(info["values"])),
            }
            text = json.dumps(data, indent=4) + "\n"
            path.write_text(text, encoding="utf-8")
            info["_path"] = path

    def _save_xlsx(self) -> None:
        ws = self._wb["Biomes"]
        # Update existing rows in place; track which biomes we've touched.
        seen: set[str] = set()
        for row in ws.iter_rows(min_row=2):
            cell_id = row[0].value
            if cell_id is None:
                continue
            raw_id = str(cell_id).strip()
            normalized = normalize_biome_id(raw_id, set(self.biome_files))
            seen.add(normalized)
            new_tier = self.tier_map.get(normalized, "")
            old_tier = (str(row[4].value).strip() if row[4].value else "")
            if new_tier != old_tier:
                row[4].value = new_tier or None
        # Append rows for biomes that have a tier assignment but no xlsx row
        # yet (e.g. alpha_islands). Source is inferred from the namespace;
        # climate/terrain/rarity columns are left blank for the user to fill in.
        for biome_id, tier in self.tier_map.items():
            if not tier or biome_id in seen:
                continue
            namespace = biome_id.split(":", 1)[0]
            source = {"minecraft": "Vanilla", "terralith": "Terralith",
                      "wythers": "WWOO"}.get(namespace, namespace)
            ws.append([biome_id, source, None, None, tier, None, None])
        self._wb.save(TIER_MAP_XLSX)


# --- helpers --------------------------------------------------------------

def _is_empty_step(v) -> bool:
    if v is None:
        return True
    if isinstance(v, list):
        return len(v) == 0
    if isinstance(v, dict):
        if not v:
            return True
        # All op-lists empty?
        for k, val in v.items():
            if k == OP_INHERIT and val:
                return False
            if k in (OP_SET, OP_ADD, OP_REMOVE) and val:
                return False
        return True
    return False


def _gather_features(node, sink: set[str]) -> None:
    if isinstance(node, list):
        for x in node:
            if isinstance(x, str):
                sink.add(x)
    elif isinstance(node, dict):
        for v in node.values():
            _gather_features(v, sink)


def _read_header(path: Path) -> str:
    """Capture leading comment block so save_all() can re-emit it verbatim."""
    if not path.exists():
        return ""
    lines: list[str] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.startswith("#") or line.strip() == "":
            lines.append(line)
        else:
            break
    return "\n".join(lines).rstrip()


# --- GUI ------------------------------------------------------------------

class GuiApp:
    SECTION_TIER = "tier"
    SECTION_TAG = "tag"

    def __init__(self, root: tk.Tk, model: DataModel) -> None:
        self.root = root
        self.model = model
        self._dirty = False
        self._build()
        self._refresh_left()

    def _build(self) -> None:
        self.root.title("Endeavour Ore Tier Manager")
        self.root.geometry("1200x720")

        # Menu bar
        menubar = tk.Menu(self.root)
        filem = tk.Menu(menubar, tearoff=0)
        filem.add_command(label="Save all (YAML + xlsx + tag JSONs)",
                          command=self._save, accelerator="Ctrl+S")
        filem.add_command(label="Reload from disk (discard edits)",
                          command=self._reload)
        filem.add_separator()
        filem.add_command(label="Apply (regenerate biome JSONs)",
                          command=self._apply)
        filem.add_command(label="Run validate.py", command=self._validate)
        filem.add_separator()
        filem.add_command(label="Quit", command=self._on_close)
        menubar.add_cascade(label="File", menu=filem)
        self.root.config(menu=menubar)
        self.root.bind("<Control-s>", lambda _e: self._save())
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

        # Toolbar
        bar = ttk.Frame(self.root, padding=(8, 6))
        bar.pack(fill="x")
        ttk.Button(bar, text="Save", command=self._save).pack(side="left")
        ttk.Button(bar, text="Reload", command=self._reload).pack(side="left", padx=(6, 0))
        ttk.Button(bar, text="Apply ->  biome JSONs",
                   command=self._apply).pack(side="left", padx=(18, 0))
        ttk.Button(bar, text="Validate", command=self._validate).pack(side="left", padx=(6, 0))
        self.status = ttk.Label(bar, text="Loaded.", anchor="e")
        self.status.pack(side="right", fill="x", expand=True)

        # Step picker (applies to ore add/remove operations)
        step_bar = ttk.Frame(self.root, padding=(8, 0, 8, 4))
        step_bar.pack(fill="x")
        ttk.Label(step_bar, text="Generation step for ore ops:").pack(side="left")
        self.step_var = tk.StringVar(value=DEFAULT_STEP)
        ttk.Combobox(step_bar, textvariable=self.step_var,
                     values=list(GENERATION_STEPS), state="readonly",
                     width=24).pack(side="left", padx=(6, 0))

        # Three-pane body
        body = ttk.PanedWindow(self.root, orient="horizontal")
        body.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        # --- left: tiers + tags
        left = ttk.Frame(body)
        body.add(left, weight=1)
        ttk.Label(left, text="Tiers & Tags", font=("", 10, "bold")).pack(anchor="w")
        self.left_tree = ttk.Treeview(left, show="tree", selectmode="browse")
        self.left_tree.pack(fill="both", expand=True, pady=(4, 4))
        self.left_tree.bind("<<TreeviewSelect>>", self._on_left_select)
        lb = ttk.Frame(left)
        lb.pack(fill="x")
        ttk.Button(lb, text="+ New tier", command=self._add_tier).pack(side="left")
        ttk.Button(lb, text="+ New tag", command=self._add_tag).pack(side="left", padx=(6, 0))
        ttk.Button(lb, text="- Delete", command=self._delete_selected).pack(side="left", padx=(6, 0))

        # --- middle: features
        middle = ttk.Frame(body)
        body.add(middle, weight=2)
        self.feat_label = ttk.Label(middle, text="Features (select tier/tag)",
                                    font=("", 10, "bold"))
        self.feat_label.pack(anchor="w")
        ttk.Label(middle, text="Filter:").pack(anchor="w", pady=(4, 0))
        self.feat_filter = tk.StringVar()
        self.feat_filter.trace_add("write", lambda *_: self._refresh_features())
        ttk.Entry(middle, textvariable=self.feat_filter).pack(fill="x")
        self.feat_list = tk.Listbox(middle, selectmode="extended",
                                    exportselection=False, activestyle="dotbox")
        self.feat_list.pack(fill="both", expand=True, pady=(4, 4))
        mb = ttk.Frame(middle)
        mb.pack(fill="x")
        ttk.Button(mb, text="+ Add ore",
                   command=self._add_feature).pack(side="left")
        ttk.Button(mb, text="- Remove",
                   command=self._remove_feature).pack(side="left", padx=(6, 0))

        # --- right: biomes
        right = ttk.Frame(body)
        body.add(right, weight=2)
        self.biome_label = ttk.Label(right, text="Biomes (select tier/tag)",
                                     font=("", 10, "bold"))
        self.biome_label.pack(anchor="w")
        ttk.Label(right, text="Filter:").pack(anchor="w", pady=(4, 0))
        self.biome_filter = tk.StringVar()
        self.biome_filter.trace_add("write", lambda *_: self._refresh_biomes())
        ttk.Entry(right, textvariable=self.biome_filter).pack(fill="x")
        self.biome_list = tk.Listbox(right, selectmode="extended",
                                     exportselection=False, activestyle="dotbox")
        self.biome_list.pack(fill="both", expand=True, pady=(4, 4))
        rb = ttk.Frame(right)
        rb.pack(fill="x")
        ttk.Button(rb, text="+ Add biome",
                   command=self._add_biome).pack(side="left")
        ttk.Button(rb, text="- Remove",
                   command=self._remove_biome).pack(side="left", padx=(6, 0))

        self._tier_iids: dict[str, str] = {}
        self._tag_iids: dict[str, str] = {}

    # --- left tree ---

    def _refresh_left(self) -> None:
        self.left_tree.delete(*self.left_tree.get_children())
        self._tier_iids.clear()
        self._tag_iids.clear()
        tiers_root = self.left_tree.insert("", "end", text="Tiers", open=True)
        for t in self.model.list_tiers():
            n_biomes = len(self.model.biomes_for_tier(t))
            iid = self.left_tree.insert(tiers_root, "end",
                                        text=f"{t}  ({n_biomes} biomes)")
            self._tier_iids[iid] = t
        tags_root = self.left_tree.insert("", "end", text="Tags", open=True)
        for tag in self.model.list_tags():
            n_biomes = len(self.model.biomes_for_tag(tag))
            iid = self.left_tree.insert(tags_root, "end",
                                        text=f"#endeavour:{tag}  ({n_biomes} biomes)")
            self._tag_iids[iid] = tag

    def _selected_section(self) -> tuple[str, str] | None:
        sel = self.left_tree.selection()
        if not sel:
            return None
        iid = sel[0]
        if iid in self._tier_iids:
            return (self.SECTION_TIER, self._tier_iids[iid])
        if iid in self._tag_iids:
            return (self.SECTION_TAG, self._tag_iids[iid])
        return None

    def _on_left_select(self, _evt=None) -> None:
        self._refresh_features()
        self._refresh_biomes()

    def _refresh_features(self) -> None:
        self.feat_list.delete(0, "end")
        sel = self._selected_section()
        step = self.step_var.get()
        if sel is None:
            self.feat_label.config(text="Features (select tier/tag)")
            return
        kind, name = sel
        if kind == self.SECTION_TIER:
            feats = self.model.features_for_tier(name, step)
            self.feat_label.config(text=f"Features in {name} @ {step}")
        else:
            feats = self.model.features_for_tag(name, step)
            self.feat_label.config(
                text=f"Features added to #endeavour:{name} @ {step}")
        flt = self.feat_filter.get().lower()
        for f in feats:
            if flt and flt not in f.lower():
                continue
            self.feat_list.insert("end", f)

    def _refresh_biomes(self) -> None:
        self.biome_list.delete(0, "end")
        sel = self._selected_section()
        if sel is None:
            self.biome_label.config(text="Biomes (select tier/tag)")
            return
        kind, name = sel
        if kind == self.SECTION_TIER:
            biomes = self.model.biomes_for_tier(name)
            self.biome_label.config(text=f"Biomes assigned to tier {name}")
        else:
            biomes = self.model.biomes_for_tag(name)
            self.biome_label.config(
                text=f"Biomes in #endeavour:{name}")
        flt = self.biome_filter.get().lower()
        for b in biomes:
            if flt and flt not in b.lower():
                continue
            self.biome_list.insert("end", b)

    # --- mutations ---

    def _add_tier(self) -> None:
        name = simpledialog.askstring("New tier",
                                      "Tier name (e.g. T3):", parent=self.root)
        if not name:
            return
        try:
            self.model.add_tier(name)
        except ValueError as e:
            messagebox.showerror("Cannot add tier", str(e), parent=self.root)
            return
        self._mark_dirty()
        self._refresh_left()

    def _add_tag(self) -> None:
        name = simpledialog.askstring(
            "New tag",
            "Tag name (without '#endeavour:' prefix):", parent=self.root)
        if not name:
            return
        try:
            self.model.add_tag(name)
        except ValueError as e:
            messagebox.showerror("Cannot add tag", str(e), parent=self.root)
            return
        self._mark_dirty()
        self._refresh_left()

    def _delete_selected(self) -> None:
        sel = self._selected_section()
        if sel is None:
            return
        kind, name = sel
        label = f"tier {name}" if kind == self.SECTION_TIER else f"tag #endeavour:{name}"
        if not messagebox.askyesno(
                "Delete?",
                f"Delete {label}?\n\n"
                f"For tiers, this only removes the template entry; biome "
                f"assignments in the xlsx are kept (you'll need to reassign "
                f"those biomes to a different tier).\n\n"
                f"For tags, this deletes the tag JSON and its entry in "
                f"biome_overrides.yaml.",
                parent=self.root):
            return
        if kind == self.SECTION_TIER:
            self.model.delete_tier(name)
        else:
            self.model.delete_tag(name)
        self._mark_dirty()
        self._refresh_left()
        self._refresh_features()
        self._refresh_biomes()

    def _add_feature(self) -> None:
        sel = self._selected_section()
        if sel is None:
            messagebox.showinfo("Pick a tier or tag",
                                "Select a tier or tag in the left pane first.",
                                parent=self.root)
            return
        kind, name = sel
        step = self.step_var.get()
        feature = _pick_from_list(
            self.root, "Add feature",
            f"Pick a feature ID to add to {kind} {name} @ {step}.\n"
            f"Or type a new ID at the bottom.",
            self.model.known_features, allow_custom=True)
        if not feature:
            return
        if kind == self.SECTION_TIER:
            self.model.add_feature_to_tier(name, feature, step)
        else:
            self.model.add_feature_to_tag(name, feature, step)
        self._mark_dirty()
        self._refresh_features()

    def _remove_feature(self) -> None:
        sel = self._selected_section()
        if sel is None:
            return
        kind, name = sel
        step = self.step_var.get()
        picked = [self.feat_list.get(i) for i in self.feat_list.curselection()]
        if not picked:
            messagebox.showinfo("Pick features",
                                "Select feature(s) in the middle list to remove.",
                                parent=self.root)
            return
        for f in picked:
            if kind == self.SECTION_TIER:
                self.model.remove_feature_from_tier(name, f, step)
            else:
                self.model.remove_feature_from_tag(name, f, step)
        self._mark_dirty()
        self._refresh_features()

    def _add_biome(self) -> None:
        sel = self._selected_section()
        if sel is None:
            messagebox.showinfo("Pick a tier or tag",
                                "Select a tier or tag in the left pane first.",
                                parent=self.root)
            return
        kind, name = sel
        if kind == self.SECTION_TIER:
            # For tier assignment, show all biomes (with their current tier
            # in the label) so the user knows what they're moving.
            options = [
                f"{b}    [now: {self.model.tier_of(b) or '-'}]"
                for b in self.model.all_biome_ids
            ]
        else:
            already = set(self.model.biomes_for_tag(name))
            options = [b for b in self.model.all_biome_ids if b not in already]
        picked = _pick_from_list(
            self.root, "Add biome",
            f"Pick biome(s) to add to {kind} {name}." +
            (" (Replaces current tier assignment.)" if kind == self.SECTION_TIER else ""),
            options, multi=True)
        if not picked:
            return
        # Strip the "[now: ...]" suffix if present
        cleaned = [p.split("    ")[0] for p in picked]
        for b in cleaned:
            if kind == self.SECTION_TIER:
                self.model.assign_biome_to_tier(b, name)
            else:
                self.model.add_biome_to_tag(b, name)
        self._mark_dirty()
        self._refresh_left()
        self._refresh_biomes()

    def _remove_biome(self) -> None:
        sel = self._selected_section()
        if sel is None:
            return
        kind, name = sel
        picked = [self.biome_list.get(i) for i in self.biome_list.curselection()]
        if not picked:
            messagebox.showinfo("Pick biomes",
                                "Select biome(s) in the right list to remove.",
                                parent=self.root)
            return
        if kind == self.SECTION_TIER:
            if not messagebox.askyesno(
                    "Clear tier assignment?",
                    f"Remove tier {name} from {len(picked)} biome(s)?\n"
                    f"They'll have no tier (resolves to empty template).",
                    parent=self.root):
                return
            for b in picked:
                self.model.assign_biome_to_tier(b, "")
        else:
            for b in picked:
                self.model.remove_biome_from_tag(b, name)
        self._mark_dirty()
        self._refresh_left()
        self._refresh_biomes()

    # --- save / reload / apply ---

    def _save(self) -> None:
        try:
            self.model.save_all()
        except Exception as e:
            messagebox.showerror("Save failed", str(e), parent=self.root)
            return
        self._dirty = False
        self.status.config(text="Saved.")
        self._refresh_left()

    def _reload(self) -> None:
        if self._dirty and not messagebox.askyesno(
                "Reload?",
                "You have unsaved changes. Discard and reload from disk?",
                parent=self.root):
            return
        self.model.load()
        self._dirty = False
        self.status.config(text="Reloaded.")
        self._refresh_left()
        self._refresh_features()
        self._refresh_biomes()

    def _apply(self) -> None:
        if self._dirty:
            if not messagebox.askyesno(
                    "Save first?",
                    "You have unsaved changes. Save before running apply.py?",
                    parent=self.root):
                return
            self._save()
        self._run_subprocess(["python", "apply.py", "--diff"])

    def _validate(self) -> None:
        self._run_subprocess(["python", "validate.py"])

    def _run_subprocess(self, cmd: list[str]) -> None:
        self.status.config(text=f"Running: {' '.join(cmd)} ...")
        self.root.update_idletasks()
        try:
            r = subprocess.run(cmd, cwd=TOOL_ROOT, capture_output=True,
                               text=True, encoding="utf-8")
        except Exception as e:
            messagebox.showerror("Run failed", str(e), parent=self.root)
            self.status.config(text="Error.")
            return
        out = (r.stdout or "") + (("\n--- stderr ---\n" + r.stderr) if r.stderr else "")
        ok = (r.returncode == 0)
        self.status.config(text=f"Exit {r.returncode}.")
        _show_text_dialog(self.root,
                          f"{cmd[1]}  (exit {r.returncode})",
                          out or "(no output)",
                          ok=ok)

    def _mark_dirty(self) -> None:
        self._dirty = True
        self.status.config(text="Modified (unsaved).")

    def _on_close(self) -> None:
        if self._dirty and not messagebox.askyesno(
                "Quit?", "You have unsaved changes. Quit anyway?",
                parent=self.root):
            return
        self.root.destroy()


# --- small reusable dialogs ----------------------------------------------

def _pick_from_list(parent: tk.Misc, title: str, prompt: str,
                    options: list[str], multi: bool = False,
                    allow_custom: bool = False) -> str | list[str] | None:
    """Modal picker with a search filter and an optional custom-text entry."""
    win = tk.Toplevel(parent)
    win.title(title)
    win.transient(parent)
    win.grab_set()
    win.geometry("520x520")

    ttk.Label(win, text=prompt, wraplength=480, justify="left").pack(
        anchor="w", padx=10, pady=(10, 4))
    ttk.Label(win, text="Filter:").pack(anchor="w", padx=10)
    filter_var = tk.StringVar()
    ttk.Entry(win, textvariable=filter_var).pack(fill="x", padx=10)
    lb = tk.Listbox(win, selectmode="extended" if multi else "browse",
                    exportselection=False, activestyle="dotbox")
    lb.pack(fill="both", expand=True, padx=10, pady=(4, 4))
    custom_var = tk.StringVar()
    if allow_custom:
        ttk.Label(win, text="Or type a custom value:").pack(anchor="w", padx=10)
        ttk.Entry(win, textvariable=custom_var).pack(fill="x", padx=10)

    def refresh():
        lb.delete(0, "end")
        f = filter_var.get().lower()
        for o in options:
            if f and f not in o.lower():
                continue
            lb.insert("end", o)

    filter_var.trace_add("write", lambda *_: refresh())
    refresh()

    result: dict = {"value": None}

    def on_ok():
        if allow_custom and custom_var.get().strip():
            result["value"] = custom_var.get().strip()
        else:
            sel = [lb.get(i) for i in lb.curselection()]
            if not sel:
                return
            result["value"] = sel if multi else sel[0]
        win.destroy()

    def on_cancel():
        win.destroy()

    btn_row = ttk.Frame(win)
    btn_row.pack(fill="x", padx=10, pady=(0, 10))
    ttk.Button(btn_row, text="OK", command=on_ok).pack(side="right")
    ttk.Button(btn_row, text="Cancel", command=on_cancel).pack(side="right", padx=(0, 6))
    lb.bind("<Double-Button-1>", lambda _e: on_ok())
    win.bind("<Return>", lambda _e: on_ok())
    win.bind("<Escape>", lambda _e: on_cancel())

    win.wait_window()
    return result["value"]


def _show_text_dialog(parent: tk.Misc, title: str, body: str, ok: bool) -> None:
    win = tk.Toplevel(parent)
    win.title(title)
    win.transient(parent)
    win.geometry("780x520")
    txt = tk.Text(win, wrap="word")
    txt.insert("1.0", body)
    txt.config(state="disabled")
    sb = ttk.Scrollbar(win, command=txt.yview)
    txt.config(yscrollcommand=sb.set)
    txt.pack(side="left", fill="both", expand=True)
    sb.pack(side="left", fill="y")
    if not ok:
        win.configure(bg="#fee")


def main() -> int:
    root = tk.Tk()
    try:
        model = DataModel()
    except Exception as e:
        messagebox.showerror("Failed to load", str(e))
        return 1
    GuiApp(root, model)
    root.mainloop()
    return 0


if __name__ == "__main__":
    sys.exit(main())
